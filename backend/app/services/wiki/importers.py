"""Импорт из .xlsx и .docx в Markdown с извлечением картинок.

Возвращают (markdown_content, [(image_name, image_bytes, content_type)]).
Картинки сохраняет endpoint, подставляя URL вида /media/{tenant_id}/wiki/{article_id}/{name}.
"""
from __future__ import annotations

import io
import re
import uuid
from typing import List, Tuple

MAX_IMPORT_BYTES = 20 * 1024 * 1024  # 20 MB верхняя граница на исходник


ImageAsset = Tuple[str, bytes, str]  # (stored_name, data, content_type)


def _safe_ext_from_content_type(ct: str) -> str:
    ct = (ct or "").lower()
    if "png" in ct:
        return "png"
    if "jpeg" in ct or "jpg" in ct:
        return "jpg"
    if "gif" in ct:
        return "gif"
    if "webp" in ct:
        return "webp"
    if "bmp" in ct:
        return "bmp"
    return "png"


def _md_escape(cell: str) -> str:
    return cell.replace("|", "\\|").replace("\n", "<br>").strip()


def xlsx_to_markdown(data: bytes) -> Tuple[str, List[ImageAsset]]:
    """Парсит xlsx-файл в Markdown. Каждый лист = раздел ##. Таблицы → md-tables.

    Картинки: openpyxl хранит их как ws._images (OpenPyXL Image objects).
    Вставляем ссылки под лист.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True)
    md_parts: List[str] = []
    images: List[ImageAsset] = []

    for ws in wb.worksheets:
        md_parts.append(f"## {ws.title}\n")

        rows_iter = ws.iter_rows(values_only=True)
        rows = [
            [("" if c is None else str(c)) for c in row]
            for row in rows_iter
            if any(c is not None and str(c).strip() for c in row)
        ]
        if rows:
            # Первая строка — заголовок таблицы (best-effort). Если все первые ячейки
            # выглядят как данные — всё равно первая идёт в шапку.
            max_cols = max(len(r) for r in rows)
            for r in rows:
                while len(r) < max_cols:
                    r.append("")

            header = rows[0]
            body = rows[1:]

            md_parts.append("| " + " | ".join(_md_escape(c) for c in header) + " |")
            md_parts.append("| " + " | ".join(["---"] * max_cols) + " |")
            for r in body:
                md_parts.append("| " + " | ".join(_md_escape(c) for c in r) + " |")
            md_parts.append("")

        # Картинки на листе
        ws_images = getattr(ws, "_images", []) or []
        for img in ws_images:
            try:
                # openpyxl.drawing.image.Image._data() возвращает bytes
                raw = img._data()  # type: ignore[attr-defined]
                ct = getattr(img, "content_type", "") or "image/png"
                ext = _safe_ext_from_content_type(ct)
                name = f"img_{uuid.uuid4().hex[:8]}.{ext}"
                images.append((name, raw, f"image/{ext if ext != 'jpg' else 'jpeg'}"))
                md_parts.append(f"![{name}](__IMG__{name})")
                md_parts.append("")
            except Exception:
                continue

        md_parts.append("")

    return "\n".join(md_parts).strip() + "\n", images


def docx_to_markdown(data: bytes) -> Tuple[str, List[ImageAsset]]:
    """Парсит docx в Markdown: заголовки, списки, параграфы, таблицы, картинки."""
    from docx import Document
    from docx.document import Document as _Doc
    from docx.oxml.ns import qn
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    doc = Document(io.BytesIO(data))
    md_parts: List[str] = []
    images: List[ImageAsset] = []

    # image relationship id -> saved name (для dedup)
    image_by_rid: dict[str, str] = {}

    def _extract_images_from_paragraph(p: Paragraph) -> List[str]:
        """Возвращает список storage names картинок в порядке появления."""
        found: List[str] = []
        blips = p._element.findall(
            ".//" + qn("a:blip")
        )
        for blip in blips:
            rid = blip.get(qn("r:embed"))
            if not rid:
                continue
            if rid in image_by_rid:
                found.append(image_by_rid[rid])
                continue
            try:
                part = doc.part.related_parts[rid]
            except KeyError:
                continue
            ct = getattr(part, "content_type", "image/png")
            ext = _safe_ext_from_content_type(ct)
            name = f"img_{uuid.uuid4().hex[:8]}.{ext}"
            images.append((name, part.blob, f"image/{ext if ext != 'jpg' else 'jpeg'}"))
            image_by_rid[rid] = name
            found.append(name)
        return found

    def _para_to_md(p: Paragraph) -> str:
        text = p.text or ""
        style = (p.style.name if p.style else "") or ""
        style_l = style.lower()

        imgs = _extract_images_from_paragraph(p)
        img_md = "\n".join(f"![{n}](__IMG__{n})" for n in imgs)

        if style_l.startswith("heading"):
            try:
                level = int(style.split()[-1])
            except ValueError:
                level = 2
            level = max(1, min(level, 6))
            base = f"{'#' * level} {text.strip()}" if text.strip() else ""
        elif style_l in ("list bullet", "list bullet 2", "list bullet 3") or _is_bullet(p):
            base = f"- {text.strip()}" if text.strip() else ""
        elif style_l.startswith("list number") or _is_numbered(p):
            base = f"1. {text.strip()}" if text.strip() else ""
        else:
            base = text.strip()

        parts = [x for x in (base, img_md) if x]
        return "\n".join(parts)

    def _is_bullet(p: Paragraph) -> bool:
        numPr = p._element.find(qn("w:pPr"))
        if numPr is None:
            return False
        numPr = numPr.find(qn("w:numPr"))
        return numPr is not None

    def _is_numbered(p: Paragraph) -> bool:
        # best-effort — та же проверка, а различие bullet/number игнорируем
        return False

    def _table_to_md(tbl: Table) -> str:
        rows = tbl.rows
        if not rows:
            return ""
        header = [_md_escape(c.text) for c in rows[0].cells]
        lines = [
            "| " + " | ".join(header) + " |",
            "| " + " | ".join(["---"] * len(header)) + " |",
        ]
        for r in rows[1:]:
            cells = [_md_escape(c.text) for c in r.cells]
            while len(cells) < len(header):
                cells.append("")
            lines.append("| " + " | ".join(cells[: len(header)]) + " |")
        return "\n".join(lines)

    # docx хранит параграфы и таблицы вперемешку — итерируем по body в порядке
    body = doc.element.body
    for child in body.iterchildren():
        tag = child.tag.split("}")[-1]
        if tag == "p":
            p = Paragraph(child, doc)
            md = _para_to_md(p)
            if md:
                md_parts.append(md)
        elif tag == "tbl":
            tbl = Table(child, doc)
            md = _table_to_md(tbl)
            if md:
                md_parts.append("")
                md_parts.append(md)
                md_parts.append("")

    return "\n\n".join(md_parts).strip() + "\n", images


def replace_image_placeholders(md: str, url_prefix: str) -> str:
    """__IMG__name → {url_prefix}/name. url_prefix БЕЗ trailing slash."""
    return re.sub(r"__IMG__([A-Za-z0-9._-]+)", lambda m: f"{url_prefix}/{m.group(1)}", md)
