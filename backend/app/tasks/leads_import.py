"""Импорт лидов (TenantLead) из XLSX или CSV.

Ожидаемые колонки (регистр не важен, порядок любой):
    name (required), contact (required), status, source, note, assignee_email

Прогресс — в Redis `import:{job_id}:progress` = "processed/total".
Ошибки — в meta задачи (первые 200).
"""
from __future__ import annotations

import csv
import io
from typing import Iterable

from openpyxl import load_workbook

from ..database import SessionLocal
from ..models import TenantLead, TenantMembership, User
from ..core.celery_app import celery_app
from ..core.redis_client import get_redis


ALLOWED_STATUSES = {"new", "contacted", "qualified", "converted", "rejected"}
REQUIRED_COLS = ("name", "contact")
OPTIONAL_COLS = ("status", "source", "note", "assignee_email")
ALL_COLS = REQUIRED_COLS + OPTIONAL_COLS


def _norm_key(k: object) -> str:
    return str(k or "").strip().lower()


def _rows_from_csv(text: str) -> list[dict[str, str]]:
    reader = csv.DictReader(io.StringIO(text))
    out: list[dict[str, str]] = []
    for row in reader:
        # Нормализуем ключи в lowercase, значения — строки.
        clean = {_norm_key(k): ("" if v is None else str(v).strip()) for k, v in row.items()}
        out.append(clean)
    return out


def _rows_from_xlsx(raw: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []

    rows_iter: Iterable = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []

    headers = [_norm_key(h) for h in header]
    out: list[dict[str, str]] = []
    for row in rows_iter:
        # Пустые строки пропускаем.
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        record: dict[str, str] = {}
        for i, cell in enumerate(row):
            if i >= len(headers):
                break
            key = headers[i]
            if not key:
                continue
            record[key] = "" if cell is None else str(cell).strip()
        out.append(record)
    return out


@celery_app.task(name="imports.import_leads", bind=True)
def import_leads(self, tenant_id: int, filename: str, raw: bytes, user_id: int) -> dict:
    """Разбирает XLSX/CSV и создаёт TenantLead для tenant_id.

    filename нужен для определения формата (.csv vs .xlsx).
    """
    ext = (filename or "").lower().rsplit(".", 1)[-1]
    rows: list[dict[str, str]] = []
    try:
        if ext == "csv":
            try:
                text = raw.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = raw.decode("cp1251")
            rows = _rows_from_csv(text)
        else:
            rows = _rows_from_xlsx(raw)
    except Exception as exc:
        return {
            "tenant_id": tenant_id,
            "total": 0,
            "created": 0,
            "errors": [{"row": 0, "name": "", "error": f"Не удалось прочитать файл: {exc}"}],
            "error_count": 1,
        }

    total = len(rows)
    redis = get_redis()
    progress_key = f"import:{self.request.id}:progress"

    def set_progress(processed: int) -> None:
        try:
            redis.set(progress_key, f"{processed}/{total}", ex=3600)
        except Exception:
            pass

    set_progress(0)

    created = 0
    errors: list[dict] = []

    with SessionLocal() as db:
        member_emails = {
            u.email.lower(): u.id
            for u in db.query(User)
            .join(TenantMembership, TenantMembership.user_id == User.id)
            .filter(TenantMembership.tenant_id == tenant_id)
            .all()
        }

        for i, row in enumerate(rows, start=1):
            try:
                name = (row.get("name") or "").strip()
                if not name:
                    raise ValueError("name обязателен")

                contact = (row.get("contact") or "").strip()
                if not contact:
                    raise ValueError("contact обязателен")

                status = (row.get("status") or "new").strip().lower() or "new"
                if status not in ALLOWED_STATUSES:
                    raise ValueError(f"недопустимый status: {status}")

                source = (row.get("source") or "import").strip()[:50] or "import"

                assignee_email = (row.get("assignee_email") or "").strip().lower()
                assignee_id = None
                if assignee_email:
                    assignee_id = member_emails.get(assignee_email)
                    if not assignee_id:
                        raise ValueError(f"пользователь {assignee_email} не в компании")

                # Всё, что не в известных колонках — уходит в custom_fields.
                custom: dict[str, str] = {
                    k: v for k, v in row.items()
                    if k and k not in ALL_COLS and v
                }

                lead = TenantLead(
                    tenant_id=tenant_id,
                    form_id=None,
                    name=name[:200],
                    contact=contact[:255],
                    custom_fields=custom,
                    note=(row.get("note") or "").strip() or None,
                    status=status,
                    source=source,
                    assignee_id=assignee_id,
                )
                db.add(lead)
                created += 1
            except Exception as exc:
                errors.append({
                    "row": i,
                    "name": row.get("name", ""),
                    "error": str(exc),
                })

            if i % 25 == 0:
                db.commit()
                set_progress(i)

        db.commit()
        set_progress(total)

    return {
        "tenant_id": tenant_id,
        "total": total,
        "created": created,
        "errors": errors[:200],
        "error_count": len(errors),
    }
